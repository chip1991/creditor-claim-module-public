from __future__ import annotations

import csv
import os
from datetime import datetime, timezone
from io import BytesIO, StringIO
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from sqlalchemy import and_, func, or_, select
from sqlalchemy.orm import Session

from app.core.auth import CurrentUser
from app.core.errors import AppError
from app.models.complaint_category import ComplaintCategoryLv1, ComplaintCategoryLv2
from app.services.audit import create_audit_log


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _norm(v: str | None) -> str:
    return (v or "").strip()


def _parse_enabled(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return True
    s = str(v).strip()
    if s in {"1", "true", "True", "启用", "yes", "Y", "y"}:
        return True
    if s in {"0", "false", "False", "禁用", "no", "N", "n"}:
        return False
    return True


def lv1_list(*, db: Session, enabled: bool | None = None) -> list[dict]:
    stmt = select(ComplaintCategoryLv1)
    if enabled is not None:
        stmt = stmt.where(ComplaintCategoryLv1.is_enabled.is_(bool(enabled)))
    rows = db.execute(stmt.order_by(ComplaintCategoryLv1.order_no.asc(), ComplaintCategoryLv1.name.asc())).scalars().all()
    return [
        {
            "id": r.id,
            "name": r.name,
            "orderNo": int(r.order_no or 0),
            "isEnabled": bool(r.is_enabled),
            "createdAt": r.created_at,
            "updatedAt": r.updated_at,
        }
        for r in rows
    ]


def lv2_page(
    *,
    db: Session,
    page: int,
    size: int,
    lv1_id: str | None,
    keyword: str | None,
    enabled: bool | None,
) -> dict:
    page = max(1, int(page or 1))
    size = max(1, min(200, int(size or 20)))

    clauses = []
    if lv1_id:
        clauses.append(ComplaintCategoryLv2.lv1_id == lv1_id.strip())
    if keyword:
        k = keyword.strip()
        clauses.append(or_(ComplaintCategoryLv2.name.ilike(f"%{k}%"), ComplaintCategoryLv2.keywords.ilike(f"%{k}%")))
    if enabled is not None:
        clauses.append(ComplaintCategoryLv2.is_enabled.is_(bool(enabled)))

    stmt = select(ComplaintCategoryLv2)
    if clauses:
        stmt = stmt.where(and_(*clauses))

    total = int(db.execute(select(func.count()).select_from(stmt.subquery())).scalar() or 0)
    rows = (
        db.execute(
            stmt.order_by(ComplaintCategoryLv2.order_no.asc(), ComplaintCategoryLv2.name.asc())
            .offset((page - 1) * size)
            .limit(size)
        )
        .scalars()
        .all()
    )

    records = [
        {
            "id": r.id,
            "lv1Id": r.lv1_id,
            "name": r.name,
            "orderNo": int(r.order_no or 0),
            "isEnabled": bool(r.is_enabled),
            "keywords": r.keywords,
            "createdAt": r.created_at,
            "updatedAt": r.updated_at,
        }
        for r in rows
    ]
    return {"total": total, "records": records, "items": records}


def lv1_create(*, db: Session, payload: dict, actor: CurrentUser) -> ComplaintCategoryLv1:
    name = _norm(payload.get("name"))
    if not name:
        raise AppError(code="VALIDATION_ERROR", msg="一级分类名称不能为空", status_code=422)

    existed = db.execute(select(func.count()).select_from(ComplaintCategoryLv1).where(ComplaintCategoryLv1.name == name)).scalar()
    if int(existed or 0) > 0:
        raise AppError(code="DUPLICATE", msg="一级分类名称已存在", status_code=400)

    row = ComplaintCategoryLv1(
        id=uuid4().hex,
        name=name,
        order_no=int(payload.get("orderNo") or 0),
        is_enabled=bool(payload.get("isEnabled", True)),
        created_by=actor.id,
        updated_by=actor.id,
    )
    db.add(row)
    create_audit_log(
        db,
        entity_type="complaint_category_lv1",
        entity_id=row.id,
        action="CATEGORY_LV1_CREATE",
        actor=actor,
        after={"name": row.name, "orderNo": row.order_no, "isEnabled": row.is_enabled},
        reason="创建投诉一级分类",
    )
    return row


def lv1_update(*, db: Session, lv1_id: str, payload: dict, actor: CurrentUser) -> ComplaintCategoryLv1:
    row = db.get(ComplaintCategoryLv1, lv1_id)
    if not row:
        raise AppError(code="NOT_FOUND", msg="一级分类不存在", status_code=404)

    before = {"name": row.name, "orderNo": int(row.order_no or 0), "isEnabled": bool(row.is_enabled)}

    if payload.get("name") is not None:
        name = _norm(payload.get("name"))
        if not name:
            raise AppError(code="VALIDATION_ERROR", msg="一级分类名称不能为空", status_code=422)
        if name != row.name:
            existed = db.execute(select(func.count()).select_from(ComplaintCategoryLv1).where(ComplaintCategoryLv1.name == name)).scalar()
            if int(existed or 0) > 0:
                raise AppError(code="DUPLICATE", msg="一级分类名称已存在", status_code=400)
        row.name = name
    if payload.get("orderNo") is not None:
        row.order_no = int(payload.get("orderNo") or 0)
    if payload.get("isEnabled") is not None:
        row.is_enabled = bool(payload.get("isEnabled"))

    row.updated_by = actor.id
    create_audit_log(
        db,
        entity_type="complaint_category_lv1",
        entity_id=row.id,
        action="CATEGORY_LV1_UPDATE",
        actor=actor,
        before=before,
        after={"name": row.name, "orderNo": row.order_no, "isEnabled": row.is_enabled},
        reason="更新投诉一级分类",
    )
    return row


def lv1_toggle(*, db: Session, lv1_id: str, is_enabled: bool, actor: CurrentUser) -> ComplaintCategoryLv1:
    row = db.get(ComplaintCategoryLv1, lv1_id)
    if not row:
        raise AppError(code="NOT_FOUND", msg="一级分类不存在", status_code=404)
    before = {"isEnabled": bool(row.is_enabled)}
    row.is_enabled = bool(is_enabled)
    row.updated_by = actor.id
    create_audit_log(
        db,
        entity_type="complaint_category_lv1",
        entity_id=row.id,
        action="CATEGORY_LV1_TOGGLE",
        actor=actor,
        before=before,
        after={"isEnabled": bool(row.is_enabled)},
        reason="启停投诉一级分类",
    )
    return row


def lv1_delete(*, db: Session, lv1_id: str, actor: CurrentUser) -> None:
    row = db.get(ComplaintCategoryLv1, lv1_id)
    if not row:
        return
    before = {"name": row.name}
    db.delete(row)
    create_audit_log(
        db,
        entity_type="complaint_category_lv1",
        entity_id=lv1_id,
        action="CATEGORY_LV1_DELETE",
        actor=actor,
        before=before,
        reason="删除投诉一级分类",
    )


def lv2_create(*, db: Session, payload: dict, actor: CurrentUser) -> ComplaintCategoryLv2:
    lv1_id = _norm(payload.get("lv1Id"))
    name = _norm(payload.get("name"))
    if not lv1_id or not name:
        raise AppError(code="VALIDATION_ERROR", msg="二级分类名称与所属一级分类为必填", status_code=422)

    existed = (
        db.execute(
            select(func.count())
            .select_from(ComplaintCategoryLv2)
            .where(ComplaintCategoryLv2.lv1_id == lv1_id, ComplaintCategoryLv2.name == name)
        ).scalar()
        or 0
    )
    if int(existed) > 0:
        raise AppError(code="DUPLICATE", msg="二级分类名称已存在", status_code=400)

    row = ComplaintCategoryLv2(
        id=uuid4().hex,
        lv1_id=lv1_id,
        name=name,
        order_no=int(payload.get("orderNo") or 0),
        is_enabled=bool(payload.get("isEnabled", True)),
        keywords=_norm(payload.get("keywords")) or None,
        created_by=actor.id,
        updated_by=actor.id,
    )
    db.add(row)
    create_audit_log(
        db,
        entity_type="complaint_category_lv2",
        entity_id=row.id,
        action="CATEGORY_LV2_CREATE",
        actor=actor,
        after={"lv1Id": row.lv1_id, "name": row.name, "orderNo": row.order_no, "isEnabled": row.is_enabled, "keywords": row.keywords},
        reason="创建投诉二级分类",
    )
    return row


def lv2_update(*, db: Session, lv2_id: str, payload: dict, actor: CurrentUser) -> ComplaintCategoryLv2:
    row = db.get(ComplaintCategoryLv2, lv2_id)
    if not row:
        raise AppError(code="NOT_FOUND", msg="二级分类不存在", status_code=404)

    before = {"lv1Id": row.lv1_id, "name": row.name, "orderNo": int(row.order_no or 0), "isEnabled": bool(row.is_enabled), "keywords": row.keywords}

    if payload.get("lv1Id") is not None:
        row.lv1_id = _norm(payload.get("lv1Id")) or row.lv1_id
    if payload.get("name") is not None:
        name = _norm(payload.get("name"))
        if not name:
            raise AppError(code="VALIDATION_ERROR", msg="二级分类名称不能为空", status_code=422)
        if name != row.name or row.lv1_id != before["lv1Id"]:
            existed = (
                db.execute(
                    select(func.count())
                    .select_from(ComplaintCategoryLv2)
                    .where(ComplaintCategoryLv2.lv1_id == row.lv1_id, ComplaintCategoryLv2.name == name, ComplaintCategoryLv2.id != row.id)
                ).scalar()
                or 0
            )
            if int(existed) > 0:
                raise AppError(code="DUPLICATE", msg="二级分类名称已存在", status_code=400)
        row.name = name
    if payload.get("orderNo") is not None:
        row.order_no = int(payload.get("orderNo") or 0)
    if payload.get("isEnabled") is not None:
        row.is_enabled = bool(payload.get("isEnabled"))
    if payload.get("keywords") is not None:
        row.keywords = _norm(payload.get("keywords")) or None

    row.updated_by = actor.id
    create_audit_log(
        db,
        entity_type="complaint_category_lv2",
        entity_id=row.id,
        action="CATEGORY_LV2_UPDATE",
        actor=actor,
        before=before,
        after={"lv1Id": row.lv1_id, "name": row.name, "orderNo": row.order_no, "isEnabled": row.is_enabled, "keywords": row.keywords},
        reason="更新投诉二级分类",
    )
    return row


def lv2_toggle(*, db: Session, lv2_id: str, is_enabled: bool, actor: CurrentUser) -> ComplaintCategoryLv2:
    row = db.get(ComplaintCategoryLv2, lv2_id)
    if not row:
        raise AppError(code="NOT_FOUND", msg="二级分类不存在", status_code=404)
    before = {"isEnabled": bool(row.is_enabled)}
    row.is_enabled = bool(is_enabled)
    row.updated_by = actor.id
    create_audit_log(
        db,
        entity_type="complaint_category_lv2",
        entity_id=row.id,
        action="CATEGORY_LV2_TOGGLE",
        actor=actor,
        before=before,
        after={"isEnabled": bool(row.is_enabled)},
        reason="启停投诉二级分类",
    )
    return row


def lv2_delete(*, db: Session, lv2_id: str, actor: CurrentUser) -> None:
    row = db.get(ComplaintCategoryLv2, lv2_id)
    if not row:
        return
    before = {"lv1Id": row.lv1_id, "name": row.name}
    db.delete(row)
    create_audit_log(
        db,
        entity_type="complaint_category_lv2",
        entity_id=lv2_id,
        action="CATEGORY_LV2_DELETE",
        actor=actor,
        before=before,
        reason="删除投诉二级分类",
    )


def _read_rows(filename: str, content: bytes) -> list[dict]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(StringIO(text))
        return [dict(r) for r in reader]

    wb = load_workbook(BytesIO(content))
    ws = wb.active
    headers = [str(c.value or "").strip() for c in ws[1]]
    rows: list[dict] = []
    for i in range(2, ws.max_row + 1):
        values = [ws.cell(row=i, column=j + 1).value for j in range(len(headers))]
        rows.append({headers[j]: values[j] for j in range(len(headers))})
    return rows


def import_xlsx_or_csv(*, db: Session, filename: str, content: bytes, actor: CurrentUser) -> dict:
    rows = _read_rows(filename, content)
    total = len(rows)
    success = 0
    failed = 0
    errors: list[dict] = []

    def get_lv1_id(name: str) -> str:
        existed = db.execute(select(ComplaintCategoryLv1).where(ComplaintCategoryLv1.name == name).limit(1)).scalars().first()
        if existed:
            return existed.id
        row = ComplaintCategoryLv1(id=uuid4().hex, name=name, order_no=0, is_enabled=True, created_by=actor.id, updated_by=actor.id)
        db.add(row)
        create_audit_log(
            db,
            entity_type="complaint_category_lv1",
            entity_id=row.id,
            action="CATEGORY_LV1_IMPORT_CREATE",
            actor=actor,
            after={"name": name},
            reason="导入新增一级分类",
        )
        return row.id

    for idx, r in enumerate(rows, start=2):
        try:
            lv1_name = _norm(r.get("一级分类") or r.get("lv1") or r.get("categoryLv1"))
            lv2_name = _norm(r.get("二级分类") or r.get("lv2") or r.get("categoryLv2"))
            keywords = _norm(r.get("关键词") or r.get("keywords")) or None
            enabled = _parse_enabled(r.get("启用") if "启用" in r else r.get("isEnabled"))
            order_no = int(r.get("排序") or r.get("orderNo") or 0)

            if not lv1_name or not lv2_name:
                raise AppError(code="VALIDATION_ERROR", msg="一级分类/二级分类不能为空", status_code=422)

            lv1_id = get_lv1_id(lv1_name)
            existed = (
                db.execute(
                    select(ComplaintCategoryLv2)
                    .where(ComplaintCategoryLv2.lv1_id == lv1_id, ComplaintCategoryLv2.name == lv2_name)
                    .limit(1)
                )
                .scalars()
                .first()
            )
            if existed:
                before = {"keywords": existed.keywords, "isEnabled": bool(existed.is_enabled), "orderNo": int(existed.order_no or 0)}
                existed.keywords = keywords
                existed.is_enabled = enabled
                existed.order_no = order_no
                existed.updated_by = actor.id
                create_audit_log(
                    db,
                    entity_type="complaint_category_lv2",
                    entity_id=existed.id,
                    action="CATEGORY_LV2_IMPORT_UPDATE",
                    actor=actor,
                    before=before,
                    after={"keywords": keywords, "isEnabled": enabled, "orderNo": order_no},
                    reason="导入更新二级分类",
                )
            else:
                row = ComplaintCategoryLv2(
                    id=uuid4().hex,
                    lv1_id=lv1_id,
                    name=lv2_name,
                    order_no=order_no,
                    is_enabled=enabled,
                    keywords=keywords,
                    created_by=actor.id,
                    updated_by=actor.id,
                )
                db.add(row)
                create_audit_log(
                    db,
                    entity_type="complaint_category_lv2",
                    entity_id=row.id,
                    action="CATEGORY_LV2_IMPORT_CREATE",
                    actor=actor,
                    after={"lv1Id": lv1_id, "name": lv2_name, "keywords": keywords, "isEnabled": enabled, "orderNo": order_no},
                    reason="导入新增二级分类",
                )

            success += 1
        except Exception as e:
            failed += 1
            msg = str(e)
            if isinstance(e, AppError):
                msg = e.msg
            errors.append({"row": idx, "error": msg})

    return {"total": total, "success": success, "failed": failed, "errors": errors}


def export_xlsx(*, db: Session) -> str:
    lv1_rows = db.execute(select(ComplaintCategoryLv1).order_by(ComplaintCategoryLv1.order_no.asc(), ComplaintCategoryLv1.name.asc())).scalars().all()
    lv1_map = {r.id: r.name for r in lv1_rows}
    lv2_rows = db.execute(select(ComplaintCategoryLv2).order_by(ComplaintCategoryLv2.lv1_id.asc(), ComplaintCategoryLv2.order_no.asc())).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "complaint_category"
    ws.append(["一级分类", "二级分类", "关键词", "启用", "排序"])
    for r in lv2_rows:
        ws.append(
            [
                lv1_map.get(r.lv1_id) or r.lv1_id,
                r.name,
                r.keywords or "",
                "启用" if r.is_enabled else "禁用",
                int(r.order_no or 0),
            ]
        )

    out_dir = "/workspace/backend/.data/exports"
    os.makedirs(out_dir, exist_ok=True)
    ts = _now().strftime("%Y%m%d%H%M%S")
    path = os.path.join(out_dir, f"complaint_category_{ts}.xlsx")
    wb.save(path)
    return path


def enabled_lv2_rules(*, db: Session) -> list[tuple[str, str, list[str]]]:
    rows = (
        db.execute(
            select(ComplaintCategoryLv1.name, ComplaintCategoryLv2.name, ComplaintCategoryLv2.keywords)
            .join(ComplaintCategoryLv2, ComplaintCategoryLv2.lv1_id == ComplaintCategoryLv1.id)
            .where(ComplaintCategoryLv1.is_enabled.is_(True), ComplaintCategoryLv2.is_enabled.is_(True))
            .order_by(ComplaintCategoryLv1.order_no.asc(), ComplaintCategoryLv2.order_no.asc())
        )
        .all()
    )

    def split_keywords(raw: str | None) -> list[str]:
        s = _norm(raw)
        if not s:
            return []
        for ch in ["，", ";", "；", "\n", "\t"]:
            s = s.replace(ch, ",")
        parts = [p.strip() for p in s.split(",")]
        return [p for p in parts if p]

    return [(lv1, lv2, split_keywords(kws)) for lv1, lv2, kws in rows]

