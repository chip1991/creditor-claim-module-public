from __future__ import annotations

import os
import re
import uuid
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import and_, func, or_, select
from sqlalchemy.orm import Session

from app.core.errors import AppError
from app.core.auth import CurrentUser
from app.db.session import SessionLocal
from app.models.data_center import (
    DataCleanLog,
    DataImportRowError,
    DataImportTask,
    DataLinkLog,
    DataRecord,
    DataStatus,
    DataType,
)
from app.models.task import TaskStatus
from app.models.rbac import RoleDataScope, User
from app.services.audit import create_audit_log
from app.services.data_scope import DataScopeProfile
from app.tasks.progress import update_task_progress


UPLOAD_DIR = "/workspace/backend/.data/uploads"
EXPORT_DIR = "/workspace/backend/.data/exports"


class ImportConflictStrategy:
    REJECT = "REJECT"
    OVERWRITE = "OVERWRITE"


def new_id() -> str:
    return uuid.uuid4().hex


def ensure_dirs() -> None:
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    os.makedirs(EXPORT_DIR, exist_ok=True)


def load_actor_stub(db: Session, actor_id: int | None) -> CurrentUser | None:
    if not actor_id:
        return None
    user = db.get(User, actor_id)
    if not user:
        return None
    profile = DataScopeProfile(
        scope=RoleDataScope.SELF,
        user_id=user.id,
        dept_id=user.department_id,
        custom_dept_ids=frozenset(),
    )
    return CurrentUser(
        id=user.id,
        username=user.username,
        dept_id=user.department_id,
        permission_codes=frozenset(),
        data_scope=profile,
    )


def save_upload_bytes(filename: str, content: bytes) -> str:
    ensure_dirs()
    safe_name = re.sub(r"[^a-zA-Z0-9.\-_]+", "_", filename)[:120] or "upload.xlsx"
    path = os.path.join(UPLOAD_DIR, f"{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}_{new_id()}_{safe_name}")
    with open(path, "wb") as f:
        f.write(content)
    return path


def excel_number_to_datetime(v: float) -> datetime:
    base = datetime(1899, 12, 30, tzinfo=timezone.utc)
    return base + timedelta(days=float(v))


def to_datetime(v: object) -> datetime | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.replace(tzinfo=v.tzinfo or timezone.utc)
    if isinstance(v, (int, float)):
        return excel_number_to_datetime(float(v))
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                dt = datetime.strptime(s, fmt)
                return dt.replace(tzinfo=timezone.utc)
            except Exception:
                continue
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
            return dt.replace(tzinfo=dt.tzinfo or timezone.utc)
        except Exception:
            return None
    return None


def norm_text(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


CONTROL_CHAR_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def clean_text(text: str) -> str:
    s = text.replace("_x000D_", "")
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = CONTROL_CHAR_RE.sub("", s)
    return s.strip()


def normalize_enum(v: str, *, mapping: dict[str, str], allow_unknown: bool = False) -> tuple[str | None, str | None]:
    raw = v.strip()
    if not raw:
        return None, None
    if raw in mapping:
        return mapping[raw], raw if mapping[raw] != raw else None
    for k, val in mapping.items():
        if raw.startswith(k):
            return val, raw if val != raw else None
    if allow_unknown:
        return raw, None
    return None, raw


@dataclass(frozen=True)
class ImportRowResult:
    ok: bool
    record: dict[str, Any] | None = None
    errors: list[tuple[str | None, str]] | None = None


def _require(value: str, field: str) -> list[tuple[str | None, str]]:
    if value.strip():
        return []
    return [(field, f"{field}不能为空")]


def _validate_data_type(value: str) -> tuple[str | None, list[tuple[str | None, str]]]:
    if value in {t.value for t in DataType}:
        return value, []
    return None, [("数据类型", f"数据类型不合法，需为：{', '.join(t.value for t in DataType)}")]


def _parse_sheet_rows(file_path: str) -> tuple[list[str], list[list[object]]]:
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = [norm_text(x) for x in rows[0]]
    data_rows = [list(r) for r in rows[1:]]
    return header, data_rows


def create_import_task_row(
    db: Session,
    *,
    task_id: str,
    filename: str,
    file_path: str,
    conflict_strategy: str,
    actor: CurrentUser | None,
) -> DataImportTask:
    row = DataImportTask(
        id=task_id,
        filename=filename,
        file_path=file_path,
        conflict_strategy=conflict_strategy,
        status=TaskStatus.RUNNING.value,
        created_by=actor.id if actor else None,
    )
    db.add(row)
    return row


def _build_row_dict(header: list[str], row: list[object]) -> dict[str, object]:
    d: dict[str, object] = {}
    for i, k in enumerate(header):
        if not k:
            continue
        d[k] = row[i] if i < len(row) else None
    return d


def _parse_row(row_dict: dict[str, object]) -> ImportRowResult:
    errors: list[tuple[str | None, str]] = []

    data_type_raw = norm_text(row_dict.get("数据类型"))
    data_type, type_errors = _validate_data_type(data_type_raw)
    errors.extend(type_errors)
    if not data_type:
        return ImportRowResult(ok=False, errors=errors)

    if data_type == DataType.COMPLAINT.value:
        work_order_no = norm_text(row_dict.get("关联工单号"))
        errors.extend(_require(work_order_no, "关联工单号"))
        owner_name = norm_text(row_dict.get("业主姓名")) or None
        building_room = norm_text(row_dict.get("楼栋房号")) or None
        phone = norm_text(row_dict.get("联系电话")) or None
        event_time = to_datetime(row_dict.get("通话时间"))
        if event_time is None:
            errors.append(("通话时间", "通话时间格式不合法"))
        duration_raw = norm_text(row_dict.get("通话时长（秒）"))
        duration_sec: int | None = None
        if duration_raw:
            if not duration_raw.isdigit():
                errors.append(("通话时长（秒）", "通话时长（秒）需为非负整数"))
            else:
                duration_sec = int(duration_raw)
        agent_name = norm_text(row_dict.get("接待坐席"))
        errors.extend(_require(agent_name, "接待坐席"))
        raw_text = norm_text(row_dict.get("原始数据内容"))
        errors.extend(_require(raw_text, "原始数据内容"))
        if errors:
            return ImportRowResult(ok=False, errors=errors)
        return ImportRowResult(
            ok=True,
            record={
                "data_type": data_type,
                "work_order_no": work_order_no,
                "owner_name": owner_name,
                "building_room": building_room,
                "phone": phone,
                "event_time": event_time,
                "duration_sec": duration_sec,
                "agent_name": agent_name,
                "satisfaction_score": None,
                "raw_text": raw_text,
                "raw_payload": dict(row_dict),
            },
        )

    if data_type == DataType.SATISFACTION_400.value:
        work_order_no = norm_text(row_dict.get("关联工单号")) or None
        owner_name = norm_text(row_dict.get("业主姓名")) or None
        building_room = norm_text(row_dict.get("楼栋房号")) or None
        phone = norm_text(row_dict.get("联系电话")) or None
        event_time = to_datetime(row_dict.get("回访时间"))
        if event_time is None:
            errors.append(("回访时间", "回访时间格式不合法"))
        agent_name = norm_text(row_dict.get("回访坐席"))
        errors.extend(_require(agent_name, "回访坐席"))
        score_raw = norm_text(row_dict.get("满意度评分（1-10）"))
        if not score_raw.isdigit() or not (1 <= int(score_raw) <= 10):
            errors.append(("满意度评分（1-10）", "满意度评分需为1-10整数"))
        raw_text = norm_text(row_dict.get("原始数据内容"))
        errors.extend(_require(raw_text, "原始数据内容"))
        if errors:
            return ImportRowResult(ok=False, errors=errors)
        return ImportRowResult(
            ok=True,
            record={
                "data_type": data_type,
                "work_order_no": work_order_no,
                "owner_name": owner_name,
                "building_room": building_room,
                "phone": phone,
                "event_time": event_time,
                "duration_sec": None,
                "agent_name": agent_name,
                "satisfaction_score": int(score_raw),
                "raw_text": raw_text,
                "raw_payload": dict(row_dict),
            },
        )

    work_order_no = norm_text(row_dict.get("关联工单号")) or None
    area_company = norm_text(row_dict.get("地区公司"))
    project_name = norm_text(row_dict.get("项目名称"))
    building_no = norm_text(row_dict.get("楼栋号码"))
    name = norm_text(row_dict.get("姓名"))
    phone = norm_text(row_dict.get("电话号码"))
    batch = norm_text(row_dict.get("任务批次"))
    assigner = norm_text(row_dict.get("分配人"))
    assign_time = to_datetime(row_dict.get("分配时间"))
    dial_time = to_datetime(row_dict.get("拨打时间"))
    status_text = norm_text(row_dict.get("状态"))
    biz_result = norm_text(row_dict.get("业务结果"))
    is_connected_raw = norm_text(row_dict.get("是否接通"))
    is_valid_raw = norm_text(row_dict.get("是否有效"))
    very_live = norm_text(row_dict.get("非常住可用"))

    for field, val in [
        ("地区公司", area_company),
        ("项目名称", project_name),
        ("楼栋号码", building_no),
        ("姓名", name),
        ("电话号码", phone),
        ("任务批次", batch),
        ("分配人", assigner),
        ("状态", status_text),
        ("业务结果", biz_result),
        ("是否接通", is_connected_raw),
        ("是否有效", is_valid_raw),
        ("非常住可用", very_live),
    ]:
        errors.extend(_require(val, field))

    if assign_time is None:
        errors.append(("分配时间", "分配时间格式不合法"))
    if dial_time is None:
        errors.append(("拨打时间", "拨打时间格式不合法"))

    if assign_time and dial_time and dial_time < assign_time:
        errors.append(("拨打时间", "拨打时间应大于等于分配时间"))

    is_connected, is_connected_bad = normalize_enum(is_connected_raw, mapping={"是": "是", "否": "否"})
    if is_connected is None:
        errors.append(("是否接通", f"是否接通不合法：{is_connected_bad or is_connected_raw}"))
    is_valid, is_valid_bad = normalize_enum(
        is_valid_raw,
        mapping={"有": "有", "无": "无", "无（无）": "无", "有（无）": "无", "有（删）": "有"},
    )
    if is_valid is None:
        errors.append(("是否有效", f"是否有效不合法：{is_valid_bad or is_valid_raw}"))

    raw_text = norm_text(row_dict.get("调查备注问题")) or norm_text(row_dict.get("一般问题")) or ""
    if errors:
        return ImportRowResult(ok=False, errors=errors)

    payload = dict(row_dict)
    payload["_normalized"] = {
        "是否接通": is_connected,
        "是否有效": is_valid,
    }
    return ImportRowResult(
        ok=True,
        record={
            "data_type": DataType.CALL_AUDIT.value,
            "work_order_no": work_order_no,
            "owner_name": name or None,
            "building_room": building_no or None,
            "phone": phone or None,
            "event_time": dial_time,
            "duration_sec": None,
            "agent_name": None,
            "satisfaction_score": None,
            "raw_text": raw_text,
            "raw_payload": payload,
        },
    )


def _pick_best_satisfaction(db: Session, *, work_order_no: str) -> DataRecord | None:
    candidates = (
        db.execute(
            select(DataRecord)
            .where(
                and_(
                    DataRecord.work_order_no == work_order_no,
                    DataRecord.data_type.in_([DataType.SATISFACTION_400.value, DataType.CALL_AUDIT.value]),
                )
            )
            .order_by(DataRecord.event_time.desc().nullslast(), DataRecord.created_at.desc())
        )
        .scalars()
        .all()
    )
    if not candidates:
        return None
    best_400 = [r for r in candidates if r.data_type == DataType.SATISFACTION_400.value]
    if best_400:
        return best_400[0]
    return candidates[0]


def run_import_task(*, task_id: str, actor_id: int | None) -> None:
    update_task_progress(None, task_id=task_id, status=TaskStatus.RUNNING, progress=0, message="开始导入")
    with SessionLocal() as db:
        import_task = db.get(DataImportTask, task_id)
        if import_task is None:
            update_task_progress(None, task_id=task_id, status=TaskStatus.FAILURE, progress=100, message="导入任务不存在")
            return

        actor = load_actor_stub(db, actor_id)
        try:
            header, data_rows = _parse_sheet_rows(import_task.file_path)
            if not header:
                raise AppError(code="IMPORT_EMPTY", msg="文件为空或缺少表头", status_code=400)

            total_rows = len(data_rows)
            import_task.total_rows = total_rows
            db.flush()

            in_file_complaint_keys: set[str] = set()
            success = 0
            failed = 0
            conflict = 0

            update_task_progress(None, task_id=task_id, progress=5, message=f"读取到{total_rows}行数据")

            for idx, row in enumerate(data_rows, start=2):
                row_dict = _build_row_dict(header, row)
                parsed = _parse_row(row_dict)
                if not parsed.ok or not parsed.record:
                    failed += 1
                    for field, msg in (parsed.errors or [(None, "行校验失败")]):
                        err = DataImportRowError(
                            import_task_id=task_id,
                            row_number=idx,
                            field=field,
                            message=msg,
                            raw_payload=dict(row_dict),
                        )
                        db.add(err)
                    continue

                rec = parsed.record
                if rec["data_type"] == DataType.COMPLAINT.value:
                    wo = str(rec["work_order_no"])
                    if wo in in_file_complaint_keys:
                        failed += 1
                        db.add(
                            DataImportRowError(
                                import_task_id=task_id,
                                row_number=idx,
                                field="关联工单号",
                                message="关联工单号在文件内重复",
                                raw_payload=dict(row_dict),
                            )
                        )
                        continue
                    in_file_complaint_keys.add(wo)

                    existing = (
                        db.execute(
                            select(DataRecord).where(
                                and_(DataRecord.data_type == DataType.COMPLAINT.value, DataRecord.work_order_no == wo)
                            )
                        )
                        .scalars()
                        .first()
                    )
                    if existing is not None:
                        if import_task.conflict_strategy == ImportConflictStrategy.OVERWRITE:
                            before = {
                                "work_order_no": existing.work_order_no,
                                "raw_text": existing.raw_text,
                                "raw_payload": existing.raw_payload,
                            }
                            existing.owner_name = rec["owner_name"]
                            existing.building_room = rec["building_room"]
                            existing.phone = rec["phone"]
                            existing.event_time = rec["event_time"]
                            existing.duration_sec = rec["duration_sec"]
                            existing.agent_name = rec["agent_name"]
                            existing.raw_text = rec["raw_text"]
                            existing.raw_payload = rec["raw_payload"]
                            existing.status = DataStatus.PENDING_CLEAN.value
                            existing.import_task_id = task_id
                            existing.created_by = actor_id
                            conflict += 1
                            create_audit_log(
                                db,
                                entity_type="data_record",
                                entity_id=existing.id,
                                action="DATA_IMPORT_OVERWRITE",
                                actor=actor,
                                before=before,
                                after={"work_order_no": existing.work_order_no, "raw_text": existing.raw_text, "raw_payload": existing.raw_payload},
                                reason="导入冲突覆盖",
                            )
                            success += 1
                        else:
                            failed += 1
                            db.add(
                                DataImportRowError(
                                    import_task_id=task_id,
                                    row_number=idx,
                                    field="关联工单号",
                                    message="关联工单号已存在，默认拒绝导入",
                                    raw_payload=dict(row_dict),
                                )
                            )
                        continue

                status = DataStatus.PENDING_CLEAN.value
                if rec["data_type"] == DataType.SATISFACTION_400.value and not rec.get("work_order_no"):
                    status = DataStatus.MATCH_FAILED.value

                record = DataRecord(
                    id=new_id(),
                    data_type=str(rec["data_type"]),
                    status=status,
                    work_order_no=rec.get("work_order_no"),
                    event_time=rec.get("event_time"),
                    duration_sec=rec.get("duration_sec"),
                    agent_name=rec.get("agent_name"),
                    owner_name=rec.get("owner_name"),
                    building_room=rec.get("building_room"),
                    phone=rec.get("phone"),
                    satisfaction_score=rec.get("satisfaction_score"),
                    raw_text=str(rec.get("raw_text") or ""),
                    raw_payload=rec.get("raw_payload"),
                    import_task_id=task_id,
                    created_by=actor_id,
                )
                db.add(record)
                success += 1

                if idx % 50 == 0:
                    pct = 5 + int(80 * idx / max(1, total_rows + 1))
                    update_task_progress(None, task_id=task_id, progress=pct, message=f"已处理到第{idx}行")

            import_task.success_rows = success
            import_task.failed_rows = failed
            import_task.conflict_rows = conflict
            import_task.status = TaskStatus.SUCCESS.value if failed == 0 else TaskStatus.SUCCESS.value
            db.commit()

            create_audit_log(
                db,
                entity_type="data_import_task",
                entity_id=task_id,
                action="DATA_IMPORT_FINISH",
                actor=actor,
                after={
                    "total": total_rows,
                    "success": success,
                    "failed": failed,
                    "conflict": conflict,
                },
            )
            db.commit()
            update_task_progress(
                None,
                task_id=task_id,
                status=TaskStatus.SUCCESS,
                progress=100,
                message=f"导入完成：成功{success}，失败{failed}，冲突{conflict}",
                extra={"success": success, "failed": failed, "conflict": conflict},
            )
        except AppError as e:
            import_task.status = TaskStatus.FAILURE.value
            db.commit()
            update_task_progress(None, task_id=task_id, status=TaskStatus.FAILURE, progress=100, message=e.msg)
        except Exception as e:
            import_task.status = TaskStatus.FAILURE.value
            db.commit()
            update_task_progress(None, task_id=task_id, status=TaskStatus.FAILURE, progress=100, message=f"导入失败：{e}")


def run_clean_task(*, task_id: str, record_ids: list[str] | None, actor_id: int | None) -> None:
    update_task_progress(None, task_id=task_id, status=TaskStatus.RUNNING, progress=0, message="开始清洗")
    with SessionLocal() as db:
        actor = load_actor_stub(db, actor_id)
        try:
            stmt = select(DataRecord).where(DataRecord.status == DataStatus.PENDING_CLEAN.value)
            if record_ids:
                stmt = stmt.where(DataRecord.id.in_(record_ids))
            rows = db.execute(stmt.order_by(DataRecord.created_at.asc())).scalars().all()
            total = len(rows)
            if total == 0:
                update_task_progress(None, task_id=task_id, status=TaskStatus.SUCCESS, progress=100, message="无需清洗")
                return

            for i, r in enumerate(rows, start=1):
                before = {"raw_text": r.raw_text, "raw_payload": r.raw_payload, "cleaned_text": r.cleaned_text, "cleaned_payload": r.cleaned_payload}
                cleaned = clean_text(r.raw_text)
                cleaned_payload = r.cleaned_payload or {}
                if r.data_type == DataType.CALL_AUDIT.value and isinstance(r.raw_payload, dict):
                    normalized = dict(r.raw_payload.get("_normalized") or {})
                    cleaned_payload = dict(r.raw_payload)
                    cleaned_payload["_cleaned"] = {"text": cleaned}
                    if normalized:
                        cleaned_payload["_normalized"] = normalized
                r.cleaned_text = cleaned
                r.cleaned_payload = cleaned_payload or None
                r.status = DataStatus.CLEANED.value
                db.add(
                    DataCleanLog(
                        record_id=r.id,
                        task_id=task_id,
                        operator_id=actor_id,
                        before=before,
                        after={"cleaned_text": r.cleaned_text, "cleaned_payload": r.cleaned_payload},
                        message="清洗完成",
                    )
                )
                if i % 50 == 0:
                    db.flush()
                pct = int(100 * i / max(1, total))
                update_task_progress(None, task_id=task_id, progress=pct, message=f"已清洗{i}/{total}")

            db.commit()
            create_audit_log(
                db,
                entity_type="task",
                entity_id=task_id,
                action="DATA_CLEAN_FINISH",
                actor=actor,
                after={"count": total},
            )
            db.commit()
            update_task_progress(None, task_id=task_id, status=TaskStatus.SUCCESS, progress=100, message=f"清洗完成：{total}条")
        except Exception as e:
            db.rollback()
            update_task_progress(None, task_id=task_id, status=TaskStatus.FAILURE, progress=100, message=f"清洗失败：{e}")


def run_link_task(*, task_id: str, record_ids: list[str] | None, actor_id: int | None) -> None:
    update_task_progress(None, task_id=task_id, status=TaskStatus.RUNNING, progress=0, message="开始关联")
    with SessionLocal() as db:
        actor = load_actor_stub(db, actor_id)
        try:
            stmt = select(DataRecord).where(DataRecord.data_type == DataType.COMPLAINT.value)
            if record_ids:
                stmt = stmt.where(DataRecord.id.in_(record_ids))
            else:
                stmt = stmt.where(DataRecord.status.in_([DataStatus.CLEANED.value, DataStatus.MATCH_FAILED.value, DataStatus.ANALYZED.value]))
            complaints = db.execute(stmt.order_by(DataRecord.created_at.asc())).scalars().all()
            total = len(complaints)
            if total == 0:
                update_task_progress(None, task_id=task_id, status=TaskStatus.SUCCESS, progress=100, message="无需关联")
                return

            for i, complaint in enumerate(complaints, start=1):
                wo = (complaint.work_order_no or "").strip()
                before_linked = complaint.linked_record_id
                if not wo:
                    complaint.status = DataStatus.MATCH_FAILED.value
                    complaint.linked_record_id = None
                    db.add(
                        DataLinkLog(
                            complaint_record_id=complaint.id,
                            satisfaction_record_id=None,
                            strategy="WORK_ORDER_NO",
                            status="FAILURE",
                            message="关联工单号为空",
                            task_id=task_id,
                            operator_id=actor_id,
                        )
                    )
                else:
                    best = _pick_best_satisfaction(db, work_order_no=wo)
                    if best is None:
                        complaint.status = DataStatus.MATCH_FAILED.value
                        complaint.linked_record_id = None
                        db.add(
                            DataLinkLog(
                                complaint_record_id=complaint.id,
                                satisfaction_record_id=None,
                                strategy="WORK_ORDER_NO",
                                status="FAILURE",
                                message="未找到可关联的满意度数据",
                                task_id=task_id,
                                operator_id=actor_id,
                            )
                        )
                    else:
                        complaint.status = DataStatus.LINKED.value
                        complaint.linked_record_id = best.id
                        best.linked_record_id = complaint.id
                        db.add(
                            DataLinkLog(
                                complaint_record_id=complaint.id,
                                satisfaction_record_id=best.id,
                                strategy="WORK_ORDER_NO",
                                status="SUCCESS",
                                message=f"关联成功，满意度来源：{best.data_type}",
                                task_id=task_id,
                                operator_id=actor_id,
                            )
                        )

                if before_linked != complaint.linked_record_id:
                    create_audit_log(
                        db,
                        entity_type="data_record",
                        entity_id=complaint.id,
                        action="DATA_LINK_UPDATE",
                        actor=actor,
                        before={"linked_record_id": before_linked, "status": complaint.status},
                        after={"linked_record_id": complaint.linked_record_id, "status": complaint.status},
                    )

                if i % 50 == 0:
                    db.flush()
                pct = int(100 * i / max(1, total))
                update_task_progress(None, task_id=task_id, progress=pct, message=f"已关联{i}/{total}")

            db.commit()
            create_audit_log(db, entity_type="task", entity_id=task_id, action="DATA_LINK_FINISH", actor=actor, after={"count": total})
            db.commit()
            update_task_progress(None, task_id=task_id, status=TaskStatus.SUCCESS, progress=100, message=f"关联完成：{total}条")
        except Exception as e:
            db.rollback()
            update_task_progress(None, task_id=task_id, status=TaskStatus.FAILURE, progress=100, message=f"关联失败：{e}")


def manual_fix_and_retry_link(
    db: Session,
    *,
    record_id: str,
    work_order_no: str,
    task_id: str,
    actor: CurrentUser | None,
) -> None:
    row = db.get(DataRecord, record_id)
    if row is None:
        raise AppError(code="NOT_FOUND", msg="数据不存在", status_code=404)
    before = {"work_order_no": row.work_order_no, "status": row.status}
    row.work_order_no = work_order_no.strip() or None
    row.status = DataStatus.CLEANED.value if row.status != DataStatus.PENDING_CLEAN.value else row.status
    db.flush()
    create_audit_log(
        db,
        entity_type="data_record",
        entity_id=row.id,
        action="DATA_MANUAL_FIX",
        actor=actor,
        before=before,
        after={"work_order_no": row.work_order_no, "status": row.status},
        reason="修复关联工单号",
    )
    if row.work_order_no:
        complaint = (
            db.execute(
                select(DataRecord).where(
                    and_(DataRecord.data_type == DataType.COMPLAINT.value, DataRecord.work_order_no == row.work_order_no)
                )
            )
            .scalars()
            .first()
        )
        if complaint is not None:
            best = _pick_best_satisfaction(db, work_order_no=row.work_order_no)
            if best is not None:
                complaint.status = DataStatus.LINKED.value
                complaint.linked_record_id = best.id
                best.linked_record_id = complaint.id
                db.add(
                    DataLinkLog(
                        complaint_record_id=complaint.id,
                        satisfaction_record_id=best.id,
                        strategy="MANUAL_FIX",
                        status="SUCCESS",
                        message=f"修复后关联成功，满意度来源：{best.data_type}",
                        task_id=task_id,
                        operator_id=actor.id if actor else None,
                    )
                )
            else:
                complaint.status = DataStatus.MATCH_FAILED.value
                complaint.linked_record_id = None
                db.add(
                    DataLinkLog(
                        complaint_record_id=complaint.id,
                        satisfaction_record_id=None,
                        strategy="MANUAL_FIX",
                        status="FAILURE",
                        message="修复后仍未找到满意度数据",
                        task_id=task_id,
                        operator_id=actor.id if actor else None,
                    )
                )


def query_data_page(
    db: Session,
    *,
    page: int,
    size: int,
    data_type: str | None,
    status: str | None,
    work_order_no: str | None,
    building_room: str | None,
    owner_keyword: str | None,
    start_time: datetime | None,
    end_time: datetime | None,
) -> tuple[list[DataRecord], int]:
    clauses = []
    if data_type:
        clauses.append(DataRecord.data_type == data_type)
    if status:
        clauses.append(DataRecord.status == status)
    if work_order_no:
        clauses.append(DataRecord.work_order_no.like(f"%{work_order_no.strip()}%"))
    if building_room:
        clauses.append(DataRecord.building_room.like(f"%{building_room.strip()}%"))
    if owner_keyword:
        k = owner_keyword.strip()
        clauses.append(or_(DataRecord.owner_name.like(f"%{k}%"), DataRecord.phone.like(f"%{k}%")))
    if start_time:
        clauses.append(DataRecord.created_at >= start_time)
    if end_time:
        clauses.append(DataRecord.created_at <= end_time)
    where = and_(*clauses) if clauses else None
    base = select(DataRecord)
    if where is not None:
        base = base.where(where)

    total = db.execute(select(func.count()).select_from(base.subquery())).scalar_one()
    rows = (
        db.execute(base.order_by(DataRecord.created_at.desc()).offset((page - 1) * size).limit(size))
        .scalars()
        .all()
    )
    return rows, int(total)


def export_to_excel(db: Session, *, rows: list[DataRecord]) -> str:
    ensure_dirs()
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    headers = [
        "数据ID",
        "数据类型",
        "数据处理状态",
        "关联工单号",
        "业主姓名",
        "楼栋房号",
        "联系电话",
        "通话/回访时间",
        "坐席",
        "满意度评分",
        "原始数据内容",
        "清洗后文本",
        "上传时间",
    ]
    ws.append(headers)
    for r in rows:
        ws.append(
            [
                r.id,
                r.data_type,
                r.status,
                r.work_order_no or "",
                r.owner_name or "",
                r.building_room or "",
                r.phone or "",
                r.event_time.isoformat() if getattr(r, "event_time", None) else "",
                r.agent_name or "",
                r.satisfaction_score or "",
                r.raw_text or "",
                r.cleaned_text or "",
                r.created_at.isoformat() if getattr(r, "created_at", None) else "",
            ]
        )
    file_path = os.path.join(EXPORT_DIR, f"{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}_{new_id()}.xlsx")
    wb.save(file_path)
    return file_path
