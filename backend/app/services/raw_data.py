from __future__ import annotations

import hashlib
import os
import re
import uuid
from datetime import datetime, timezone
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import and_, case, func, or_, select
from sqlalchemy.orm import Session

from app.core.auth import CurrentUser
from app.core.errors import AppError
from app.db.session import SessionLocal
from app.models.raw_data import RawDataBatch, RawDataRow, RawIssue
from app.models.system_rule import SystemRule
from app.models.task import TaskStatus
from app.models.rbac import RoleDataScope, User
from app.services.audit import create_audit_log
from app.services.data_center import to_datetime
from app.services.data_scope import DataScopeProfile
from app.tasks.progress import update_task_progress


UPLOAD_DIR = "/workspace/backend/.data/raw_uploads"


HEADERS = [
    "地区公司",
    "项目名称",
    "楼栋号码",
    "是否过保",
    "姓名",
    "电话号码",
    "性别",
    "任务批次",
    "分配人",
    "分配时间",
    "拨打时间",
    "状态",
    "业务结果",
    "是否接通",
    "是否有效",
    "非常住可用",
    "首轮评价",
    "居住情况",
    "400类",
    "一般问题",
    "管家服务",
    "安保服务",
    "环境卫生",
    "公区维修",
    "调查备注问题",
]


DEFAULT_SCORE_MAPPING = {"满意": 100, "一般": 80, "不满意": 60}


def new_id() -> str:
    return uuid.uuid4().hex


def _ensure_dirs() -> None:
    os.makedirs(UPLOAD_DIR, exist_ok=True)


def _safe_filename(name: str) -> str:
    return re.sub(r"[^a-zA-Z0-9.\-_]+", "_", name)[:120] or "upload.xlsx"


def save_upload_bytes(filename: str, content: bytes) -> tuple[str, str]:
    _ensure_dirs()
    file_hash = hashlib.sha256(content).hexdigest()
    safe_name = _safe_filename(filename)
    path = os.path.join(UPLOAD_DIR, f"{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}_{new_id()}_{safe_name}")
    with open(path, "wb") as f:
        f.write(content)
    return path, file_hash


def safe_delete_upload_file(file_path: str | None) -> bool:
    if not file_path:
        return False
    try:
        real_base = os.path.realpath(UPLOAD_DIR)
        real_target = os.path.realpath(file_path)
        if not real_target.startswith(real_base + os.sep):
            return False
        if not os.path.isfile(real_target):
            return False
        os.remove(real_target)
        return True
    except Exception:
        return False


def load_actor_stub(db: Session, actor_id: int | None) -> CurrentUser | None:
    if not actor_id:
        return None
    user = db.get(User, actor_id)
    if not user:
        return None
    profile = DataScopeProfile(
        scope=RoleDataScope.SELF,
        user_id=user.id,
        dept_id=user.department_id,
        custom_dept_ids=frozenset(),
    )
    return CurrentUser(
        id=user.id,
        username=user.username,
        dept_id=user.department_id,
        permission_codes=frozenset(),
        menu_ids=frozenset(),
        data_scope=profile,
    )


def _normalize_payload(payload: dict[str, object]) -> dict[str, object]:
    out: dict[str, object] = {}
    for k, v in payload.items():
        if isinstance(v, datetime):
            out[k] = v.isoformat()
        else:
            out[k] = v
    return out


def create_raw_batch_row(
    db: Session,
    *,
    batch_id: str,
    filename: str,
    file_path: str,
    file_hash: str | None,
    sheet_name: str,
    actor: CurrentUser | None,
) -> RawDataBatch:
    row = RawDataBatch(
        id=batch_id,
        filename=filename,
        file_path=file_path,
        file_hash=file_hash,
        sheet_name=sheet_name,
        status=TaskStatus.RUNNING.value,
        created_by=actor.id if actor else None,
    )
    db.add(row)
    return row


def _extract_issues(text: str | None) -> list[str]:
    if not text:
        return []
    parts = [x.strip() for x in str(text).replace("\r\n", "\n").split("\n")]
    return [p for p in parts if p]


def get_score_mapping(db: Session) -> dict[str, int]:
    row = db.get(SystemRule, "raw_score_mapping_v1")
    if not row or not isinstance(row.value, dict):
        return dict(DEFAULT_SCORE_MAPPING)
    out: dict[str, int] = {}
    for k, v in row.value.items():
        if k is None:
            continue
        try:
            out[str(k)] = int(v)
        except Exception:
            continue
    return out or dict(DEFAULT_SCORE_MAPPING)


def set_score_mapping(db: Session, mapping: dict) -> dict[str, int]:
    normalized: dict[str, int] = {}
    for k, v in mapping.items():
        if k is None:
            continue
        normalized[str(k)] = int(v)
    row = db.get(SystemRule, "raw_score_mapping_v1")
    if row is None:
        row = SystemRule(key="raw_score_mapping_v1", value=normalized)
        db.add(row)
    else:
        row.value = normalized
    return normalized


def _score_case(field, mapping: dict[str, int]):
    whens = [(field == k, v) for k, v in mapping.items()]
    return case(*whens, else_=None)


def run_raw_import_task(*, task_id: str, batch_id: str, actor_id: int | None) -> None:
    update_task_progress(None, task_id=task_id, status=TaskStatus.RUNNING, progress=0, message="开始导入")
    with SessionLocal() as db:
        batch = db.get(RawDataBatch, batch_id)
        if batch is None:
            update_task_progress(None, task_id=task_id, status=TaskStatus.FAILURE, progress=100, message="批次不存在")
            return
        actor = load_actor_stub(db, actor_id)
        try:
            wb = load_workbook(batch.file_path, data_only=True, read_only=True)
            if batch.sheet_name not in wb.sheetnames:
                raise AppError(code="RAW_SHEET_NOT_FOUND", msg=f"未找到工作表：{batch.sheet_name}", status_code=400)
            ws = wb[batch.sheet_name]

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                raise AppError(code="RAW_EMPTY", msg="文件为空或缺少表头", status_code=400)
            header_list = [str(x).strip() if x is not None else "" for x in header_row]
            if header_list[: len(HEADERS)] != HEADERS:
                raise AppError(code="RAW_HEADER_MISMATCH", msg="Excel 表头不匹配，无法导入", status_code=400)

            total = 0
            success = 0
            failed = 0
            update_task_progress(None, task_id=task_id, progress=3, message="开始读取行数据")

            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(v in (None, "") for v in row):
                    continue
                total += 1
                try:
                    row_payload = {HEADERS[i]: row[i] if i < len(row) else None for i in range(len(HEADERS))}
                    assigned_at = to_datetime(row_payload.get("分配时间"))
                    dialed_at = to_datetime(row_payload.get("拨打时间"))
                    record = RawDataRow(
                        batch_id=batch_id,
                        row_no=idx,
                        region_company=str(row_payload.get("地区公司") or "").strip() or None,
                        project_name=str(row_payload.get("项目名称") or "").strip() or None,
                        building_no=str(row_payload.get("楼栋号码") or "").strip() or None,
                        warranty_date=str(row_payload.get("是否过保") or "").strip() or None,
                        owner_name=str(row_payload.get("姓名") or "").strip() or None,
                        phone=str(row_payload.get("电话号码") or "").strip() or None,
                        gender=str(row_payload.get("性别") or "").strip() or None,
                        task_batch=str(row_payload.get("任务批次") or "").strip() or None,
                        assigned_by=str(row_payload.get("分配人") or "").strip() or None,
                        assigned_at=assigned_at,
                        dialed_at=dialed_at,
                        status=str(row_payload.get("状态") or "").strip() or None,
                        biz_result=str(row_payload.get("业务结果") or "").strip() or None,
                        is_connected=str(row_payload.get("是否接通") or "").strip() or None,
                        is_valid=str(row_payload.get("是否有效") or "").strip() or None,
                        non_resident_usable=str(row_payload.get("非常住可用") or "").strip() or None,
                        first_rating=str(row_payload.get("首轮评价") or "").strip() or None,
                        living_status=str(row_payload.get("居住情况") or "").strip() or None,
                        call400_category=str(row_payload.get("400类") or "").strip() or None,
                        general_issue=str(row_payload.get("一般问题") or "").strip() or None,
                        butler_service=str(row_payload.get("管家服务") or "").strip() or None,
                        security_service=str(row_payload.get("安保服务") or "").strip() or None,
                        env_hygiene=str(row_payload.get("环境卫生") or "").strip() or None,
                        public_repair=str(row_payload.get("公区维修") or "").strip() or None,
                        remark_issue=str(row_payload.get("调查备注问题") or "").strip() or None,
                        raw_payload=_normalize_payload({k: v for k, v in row_payload.items()}),
                    )
                    db.add(record)
                    db.flush()

                    for issue_text in _extract_issues(record.general_issue):
                        db.add(
                            RawIssue(
                                batch_id=batch_id,
                                row_id=record.id,
                                source_field="一般问题",
                                issue_text=issue_text,
                                region_company=record.region_company,
                                project_name=record.project_name,
                                building_no=record.building_no,
                                task_batch=record.task_batch,
                                dialed_at=record.dialed_at,
                            )
                        )

                    success += 1
                except Exception:
                    failed += 1

                if total % 200 == 0:
                    db.flush()
                    db.commit()
                    pct = min(95, 3 + int(90 * total / max(1, ws.max_row - 1)))
                    update_task_progress(None, task_id=task_id, progress=pct, message=f"已导入{total}行")

            batch.total_rows = total
            batch.success_rows = success
            batch.failed_rows = failed
            batch.status = TaskStatus.SUCCESS.value if success > 0 else TaskStatus.FAILURE.value
            db.commit()

            create_audit_log(
                db,
                entity_type="raw_data_batch",
                entity_id=batch_id,
                action="RAW_IMPORT_FINISH",
                actor=actor,
                after={"total": total, "success": success, "failed": failed, "filename": batch.filename},
            )
            db.commit()
            update_task_progress(
                None,
                task_id=task_id,
                status=TaskStatus.SUCCESS if success > 0 else TaskStatus.FAILURE,
                progress=100,
                message=f"导入完成：成功{success}，失败{failed}",
                extra={"batchId": batch_id, "success": success, "failed": failed},
            )
        except AppError as e:
            batch.status = TaskStatus.FAILURE.value
            db.commit()
            create_audit_log(
                db,
                entity_type="raw_data_batch",
                entity_id=batch_id,
                action="RAW_IMPORT_FAIL",
                actor=actor,
                after={"error": e.msg, "filename": batch.filename},
            )
            db.commit()
            update_task_progress(None, task_id=task_id, status=TaskStatus.FAILURE, progress=100, message=e.msg)
        except Exception as e:
            batch.status = TaskStatus.FAILURE.value
            db.commit()
            create_audit_log(
                db,
                entity_type="raw_data_batch",
                entity_id=batch_id,
                action="RAW_IMPORT_FAIL",
                actor=actor,
                after={"error": str(e), "filename": batch.filename},
            )
            db.commit()
            update_task_progress(None, task_id=task_id, status=TaskStatus.FAILURE, progress=100, message=f"导入失败：{e}")


def query_raw_batch_page(
    db: Session,
    *,
    page: int,
    size: int,
    keyword: str | None,
    status: str | None,
    start_time: datetime | None,
    end_time: datetime | None,
) -> tuple[list[RawDataBatch], int]:
    base = select(RawDataBatch)
    clauses = []
    if keyword:
        kw = f"%{keyword.strip()}%"
        clauses.append(RawDataBatch.filename.ilike(kw))
    if status:
        clauses.append(RawDataBatch.status == status)
    if start_time:
        clauses.append(RawDataBatch.created_at >= start_time)
    if end_time:
        clauses.append(RawDataBatch.created_at <= end_time)
    if clauses:
        base = base.where(and_(*clauses))
    total = int(db.execute(select(func.count()).select_from(base.subquery())).scalar() or 0)
    rows = (
        db.execute(base.order_by(RawDataBatch.created_at.desc()).offset((page - 1) * size).limit(size))
        .scalars()
        .all()
    )
    return rows, total


def query_raw_row_page(
    db: Session,
    *,
    batch_id: str,
    page: int,
    size: int,
    region_company: str | None,
    project_name: str | None,
    task_batch: str | None,
    status: str | None,
    biz_result: str | None,
    is_connected: str | None,
    is_valid: str | None,
    first_rating: str | None,
    keyword: str | None,
) -> tuple[list[RawDataRow], int]:
    base = select(RawDataRow).where(RawDataRow.batch_id == batch_id)
    clauses = []
    if region_company:
        clauses.append(RawDataRow.region_company == region_company)
    if project_name:
        clauses.append(RawDataRow.project_name == project_name)
    if task_batch:
        clauses.append(RawDataRow.task_batch == task_batch)
    if status:
        clauses.append(RawDataRow.status == status)
    if biz_result:
        clauses.append(RawDataRow.biz_result == biz_result)
    if is_connected:
        clauses.append(RawDataRow.is_connected == is_connected)
    if is_valid:
        clauses.append(RawDataRow.is_valid == is_valid)
    if first_rating:
        clauses.append(RawDataRow.first_rating == first_rating)
    if keyword:
        kw = f"%{keyword.strip()}%"
        clauses.append(
            or_(
                RawDataRow.owner_name.ilike(kw),
                RawDataRow.phone.ilike(kw),
                RawDataRow.building_no.ilike(kw),
                RawDataRow.general_issue.ilike(kw),
                RawDataRow.remark_issue.ilike(kw),
            )
        )
    if clauses:
        base = base.where(and_(*clauses))
    total = int(db.execute(select(func.count()).select_from(base.subquery())).scalar() or 0)
    rows = (
        db.execute(base.order_by(RawDataRow.row_no.asc()).offset((page - 1) * size).limit(size))
        .scalars()
        .all()
    )
    return rows, total


def query_raw_issue_page(
    db: Session,
    *,
    page: int,
    size: int,
    batch_id: str | None,
    keyword: str | None,
    region_company: str | None,
    project_name: str | None,
    task_batch: str | None,
    source_field: str | None,
) -> tuple[list[RawIssue], int]:
    base = select(RawIssue)
    clauses = []
    if batch_id:
        clauses.append(RawIssue.batch_id == batch_id)
    if region_company:
        clauses.append(RawIssue.region_company == region_company)
    if project_name:
        clauses.append(RawIssue.project_name == project_name)
    if task_batch:
        clauses.append(RawIssue.task_batch == task_batch)
    if source_field is None:
        clauses.append(RawIssue.source_field == "一般问题")
    elif source_field == "一般问题":
        clauses.append(RawIssue.source_field == "一般问题")
    else:
        clauses.append(RawIssue.source_field == "__none__")
    if keyword:
        kw = f"%{keyword.strip()}%"
        clauses.append(RawIssue.issue_text.ilike(kw))
    if clauses:
        base = base.where(and_(*clauses))
    total = int(db.execute(select(func.count()).select_from(base.subquery())).scalar() or 0)
    rows = (
        db.execute(base.order_by(RawIssue.created_at.desc()).offset((page - 1) * size).limit(size))
        .scalars()
        .all()
    )
    return rows, total


def query_raw_score_summary(
    db: Session,
    *,
    batch_id: str | None,
    group_by: str,
    only_valid_connected: bool,
) -> list[dict[str, Any]]:
    col_map = {
        "regionCompany": "region_company",
        "projectName": "project_name",
        "taskBatch": "task_batch",
    }
    if group_by not in col_map:
        raise AppError(code="RAW_GROUP_BY_INVALID", msg="groupBy 参数不合法", status_code=400)
    group_col = getattr(RawDataRow, col_map[group_by])

    clauses = []
    if batch_id:
        clauses.append(RawDataRow.batch_id == batch_id)
    if only_valid_connected:
        clauses.append(and_(RawDataRow.is_valid == "有", RawDataRow.is_connected == "是"))

    mapping = get_score_mapping(db)
    first_case = _score_case(RawDataRow.first_rating, mapping)
    butler_case = _score_case(RawDataRow.butler_service, mapping)
    security_case = _score_case(RawDataRow.security_service, mapping)
    env_case = _score_case(RawDataRow.env_hygiene, mapping)
    public_case = _score_case(RawDataRow.public_repair, mapping)

    stmt = select(
        group_col,
        func.count().label("sample_count"),
        func.count(first_case).label("rated_count_first"),
        func.avg(first_case).label("avg_first"),
        func.count(butler_case).label("rated_count_butler"),
        func.avg(butler_case).label("avg_butler"),
        func.count(security_case).label("rated_count_security"),
        func.avg(security_case).label("avg_security"),
        func.count(env_case).label("rated_count_env"),
        func.avg(env_case).label("avg_env"),
        func.count(public_case).label("rated_count_public"),
        func.avg(public_case).label("avg_public"),
    ).select_from(RawDataRow)
    if clauses:
        stmt = stmt.where(and_(*clauses))
    stmt = stmt.group_by(group_col).order_by(func.count().desc())

    rows = db.execute(stmt).all()
    out: list[dict[str, Any]] = []
    for r in rows:
        group_val = r[0]
        out.append(
            {
                "key": str(group_val or "-"),
                "regionCompany": group_val if group_by == "regionCompany" else None,
                "projectName": group_val if group_by == "projectName" else None,
                "taskBatch": group_val if group_by == "taskBatch" else None,
                "sampleCount": int(r.sample_count or 0),
                "ratedCountFirst": int(r.rated_count_first or 0),
                "avgFirst": float(r.avg_first) if r.avg_first is not None else None,
                "ratedCountButler": int(r.rated_count_butler or 0),
                "avgButler": float(r.avg_butler) if r.avg_butler is not None else None,
                "ratedCountSecurity": int(r.rated_count_security or 0),
                "avgSecurity": float(r.avg_security) if r.avg_security is not None else None,
                "ratedCountEnv": int(r.rated_count_env or 0),
                "avgEnv": float(r.avg_env) if r.avg_env is not None else None,
                "ratedCountPublicRepair": int(r.rated_count_public or 0),
                "avgPublicRepair": float(r.avg_public) if r.avg_public is not None else None,
            }
        )
    return out
