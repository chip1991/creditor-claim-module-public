from __future__ import annotations

import csv
import os
from datetime import datetime, timezone
from io import BytesIO, StringIO
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from sqlalchemy import and_, func, or_, select
from sqlalchemy.orm import Session

from app.core.auth import CurrentUser
from app.core.errors import AppError
from app.models.root_cause_kb import RootCauseKb
from app.services.audit import create_audit_log
from app.services.analysis import _CATEGORY_LV2_BY_LV1


LEVEL_NAME_TO_CODE = {"表层问题": "surface", "直接原因": "direct", "深层管理根因": "deep"}
LEVEL_CODE_TO_NAME = {v: k for k, v in LEVEL_NAME_TO_CODE.items()}


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _norm(s: str | None) -> str:
    return (s or "").strip()


def _parse_level(v: str) -> str:
    raw = _norm(v)
    if raw in {"surface", "direct", "deep"}:
        return raw
    if raw in LEVEL_NAME_TO_CODE:
        return LEVEL_NAME_TO_CODE[raw]
    raise AppError(code="LEVEL_INVALID", msg="根因层级不合法", status_code=422)


def _parse_enabled(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return True
    s = str(v).strip()
    if s in {"1", "true", "True", "启用", "yes", "Y", "y"}:
        return True
    if s in {"0", "false", "False", "禁用", "no", "N", "n"}:
        return False
    return True


def kb_categories() -> list[dict]:
    items = []
    for lv1, lv2s in _CATEGORY_LV2_BY_LV1.items():
        items.append({"categoryLv1": lv1, "categoryLv2List": list(lv2s)})
    return items


def kb_page(
    *,
    db: Session,
    page: int,
    size: int,
    keyword: str | None,
    category_lv1: str | None,
    category_lv2: str | None,
    level: str | None,
    is_enabled: bool | None,
) -> dict:
    page = max(1, int(page or 1))
    size = max(1, min(200, int(size or 20)))

    clauses = []
    if keyword:
        k = keyword.strip()
        clauses.append(or_(RootCauseKb.content.ilike(f"%{k}%"), RootCauseKb.keywords.ilike(f"%{k}%")))
    if category_lv1:
        clauses.append(RootCauseKb.category_lv1 == category_lv1.strip())
    if category_lv2:
        clauses.append(RootCauseKb.category_lv2 == category_lv2.strip())
    if level:
        clauses.append(RootCauseKb.level == _parse_level(level))
    if is_enabled is not None:
        clauses.append(RootCauseKb.is_enabled.is_(bool(is_enabled)))

    base = select(RootCauseKb)
    if clauses:
        base = base.where(and_(*clauses))

    total = int(db.execute(select(func.count()).select_from(base.subquery())).scalar() or 0)
    rows = (
        db.execute(
            base.order_by(RootCauseKb.updated_at.desc(), RootCauseKb.id.desc())
            .offset((page - 1) * size)
            .limit(size)
        )
        .scalars()
        .all()
    )

    def to_item(r: RootCauseKb) -> dict:
        return {
            "id": r.id,
            "categoryLv1": r.category_lv1,
            "categoryLv2": r.category_lv2,
            "level": r.level,
            "content": r.content,
            "keywords": r.keywords,
            "isEnabled": bool(r.is_enabled),
            "createdAt": r.created_at,
            "updatedAt": r.updated_at,
        }

    records = [to_item(r) for r in rows]
    return {"total": total, "records": records, "items": records}


def kb_create(*, db: Session, payload: dict, actor: CurrentUser) -> RootCauseKb:
    category_lv1 = _norm(payload.get("categoryLv1"))
    category_lv2 = _norm(payload.get("categoryLv2"))
    level = _parse_level(str(payload.get("level") or ""))
    content = _norm(payload.get("content"))
    keywords = _norm(payload.get("keywords")) or None
    is_enabled = bool(payload.get("isEnabled", True))

    if not category_lv1 or not category_lv2 or not content:
        raise AppError(code="VALIDATION_ERROR", msg="分类与根因内容为必填", status_code=422)

    existed = (
        db.execute(
            select(RootCauseKb)
            .where(
                RootCauseKb.category_lv2 == category_lv2,
                RootCauseKb.level == level,
                RootCauseKb.content == content,
            )
            .limit(1)
        )
        .scalars()
        .first()
    )
    if existed:
        raise AppError(code="DUPLICATE", msg="根因条目已存在", status_code=400)

    row = RootCauseKb(
        id=uuid4().hex,
        category_lv1=category_lv1,
        category_lv2=category_lv2,
        level=level,
        content=content,
        keywords=keywords,
        is_enabled=is_enabled,
        created_by=actor.id,
        updated_by=actor.id,
    )
    db.add(row)
    create_audit_log(
        db,
        entity_type="root_cause_kb",
        entity_id=row.id,
        action="ROOT_CAUSE_KB_CREATE",
        actor=actor,
        after={
            "categoryLv1": category_lv1,
            "categoryLv2": category_lv2,
            "level": level,
            "content": content,
            "keywords": keywords,
            "isEnabled": is_enabled,
        },
        reason="创建根因条目",
    )
    return row


def kb_update(*, db: Session, kb_id: str, payload: dict, actor: CurrentUser) -> RootCauseKb:
    row = db.get(RootCauseKb, kb_id)
    if not row:
        raise AppError(code="NOT_FOUND", msg="根因条目不存在", status_code=404)

    before = {
        "categoryLv1": row.category_lv1,
        "categoryLv2": row.category_lv2,
        "level": row.level,
        "content": row.content,
        "keywords": row.keywords,
        "isEnabled": bool(row.is_enabled),
    }

    if payload.get("categoryLv1") is not None:
        row.category_lv1 = _norm(payload.get("categoryLv1")) or row.category_lv1
    if payload.get("categoryLv2") is not None:
        row.category_lv2 = _norm(payload.get("categoryLv2")) or row.category_lv2
    if payload.get("level") is not None:
        row.level = _parse_level(str(payload.get("level") or row.level))
    if payload.get("content") is not None:
        row.content = _norm(payload.get("content")) or row.content
    if payload.get("keywords") is not None:
        row.keywords = _norm(payload.get("keywords")) or None
    if payload.get("isEnabled") is not None:
        row.is_enabled = bool(payload.get("isEnabled"))

    if not _norm(row.category_lv1) or not _norm(row.category_lv2) or not _norm(row.content):
        raise AppError(code="VALIDATION_ERROR", msg="分类与根因内容为必填", status_code=422)

    row.updated_by = actor.id
    create_audit_log(
        db,
        entity_type="root_cause_kb",
        entity_id=row.id,
        action="ROOT_CAUSE_KB_UPDATE",
        actor=actor,
        before=before,
        after={
            "categoryLv1": row.category_lv1,
            "categoryLv2": row.category_lv2,
            "level": row.level,
            "content": row.content,
            "keywords": row.keywords,
            "isEnabled": bool(row.is_enabled),
        },
        reason="更新根因条目",
    )
    return row


def kb_toggle(*, db: Session, kb_id: str, is_enabled: bool, actor: CurrentUser) -> RootCauseKb:
    row = db.get(RootCauseKb, kb_id)
    if not row:
        raise AppError(code="NOT_FOUND", msg="根因条目不存在", status_code=404)
    before = {"isEnabled": bool(row.is_enabled)}
    row.is_enabled = bool(is_enabled)
    row.updated_by = actor.id
    create_audit_log(
        db,
        entity_type="root_cause_kb",
        entity_id=row.id,
        action="ROOT_CAUSE_KB_TOGGLE",
        actor=actor,
        before=before,
        after={"isEnabled": bool(row.is_enabled)},
        reason="启停根因条目",
    )
    return row


def kb_delete(*, db: Session, kb_id: str, actor: CurrentUser) -> None:
    row = db.get(RootCauseKb, kb_id)
    if not row:
        return
    before = {
        "categoryLv1": row.category_lv1,
        "categoryLv2": row.category_lv2,
        "level": row.level,
        "content": row.content,
    }
    db.delete(row)
    create_audit_log(
        db,
        entity_type="root_cause_kb",
        entity_id=kb_id,
        action="ROOT_CAUSE_KB_DELETE",
        actor=actor,
        before=before,
        reason="删除根因条目",
    )


def _split_keywords(raw: str | None) -> list[str]:
    s = _norm(raw)
    if not s:
        return []
    for ch in ["，", ";", "；", "\n", "\t"]:
        s = s.replace(ch, ",")
    parts = [p.strip() for p in s.split(",")]
    return [p for p in parts if p]


def _score(text: str, keywords: list[str]) -> int:
    if not keywords:
        return 0
    t = text or ""
    return sum(t.count(k) for k in keywords if k)


def pick_root_causes_from_kb(
    *,
    db: Session,
    text: str,
    category_lv1: str,
    category_lv2: str,
    min_score: int = 1,
) -> tuple[str | None, str | None, str | None]:
    rows = (
        db.execute(
            select(RootCauseKb)
            .where(
                RootCauseKb.is_enabled.is_(True),
                RootCauseKb.category_lv2 == category_lv2,
            )
            .order_by(RootCauseKb.updated_at.desc())
        )
        .scalars()
        .all()
    )

    if not rows:
        rows = (
            db.execute(
                select(RootCauseKb)
                .where(
                    RootCauseKb.is_enabled.is_(True),
                    RootCauseKb.category_lv1 == category_lv1,
                    RootCauseKb.category_lv2 == category_lv2,
                )
                .order_by(RootCauseKb.updated_at.desc())
            )
            .scalars()
            .all()
        )

    def best(level_code: str) -> str | None:
        candidates = [r for r in rows if r.level == level_code]
        scored: list[tuple[int, RootCauseKb]] = []
        for r in candidates:
            scored.append((_score(text, _split_keywords(r.keywords)), r))
        scored.sort(key=lambda x: x[0], reverse=True)
        if not scored:
            return None
        if scored[0][0] < min_score:
            return None
        return scored[0][1].content

    return best("surface"), best("direct"), best("deep")


def kb_import_xlsx_or_csv(*, db: Session, filename: str, content: bytes, actor: CurrentUser) -> dict:
    name = (filename or "").lower()
    rows: list[dict] = []

    if name.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="ignore")
        reader = csv.DictReader(StringIO(text))
        for r in reader:
            rows.append(dict(r))
    else:
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        headers = [str(c.value or "").strip() for c in ws[1]]
        for i in range(2, ws.max_row + 1):
            values = [ws.cell(row=i, column=j + 1).value for j in range(len(headers))]
            rows.append({headers[j]: values[j] for j in range(len(headers))})

    total = len(rows)
    success = 0
    failed = 0
    errors: list[dict] = []

    for idx, r in enumerate(rows, start=2):
        try:
            category_lv1 = _norm(r.get("一级分类") or r.get("categoryLv1") or r.get("category_lv1"))
            category_lv2 = _norm(r.get("二级分类") or r.get("categoryLv2") or r.get("category_lv2"))
            level_raw = _norm(r.get("根因层级") or r.get("level"))
            content_raw = _norm(r.get("根因内容") or r.get("content"))
            keywords_raw = _norm(r.get("关键词") or r.get("keywords")) or None
            enabled_raw = r.get("启用") if "启用" in r else r.get("isEnabled")

            if not category_lv1 or not category_lv2 or not level_raw or not content_raw:
                raise AppError(code="VALIDATION_ERROR", msg="分类/层级/根因内容不能为空", status_code=422)

            level = _parse_level(level_raw)
            is_enabled = _parse_enabled(enabled_raw)

            existed = (
                db.execute(
                    select(RootCauseKb)
                    .where(
                        RootCauseKb.category_lv2 == category_lv2,
                        RootCauseKb.level == level,
                        RootCauseKb.content == content_raw,
                    )
                    .limit(1)
                )
                .scalars()
                .first()
            )
            before = None
            if existed:
                before = {
                    "keywords": existed.keywords,
                    "isEnabled": bool(existed.is_enabled),
                }
                existed.keywords = keywords_raw
                existed.is_enabled = is_enabled
                existed.updated_by = actor.id
                create_audit_log(
                    db,
                    entity_type="root_cause_kb",
                    entity_id=existed.id,
                    action="ROOT_CAUSE_KB_IMPORT_UPDATE",
                    actor=actor,
                    before=before,
                    after={"keywords": keywords_raw, "isEnabled": is_enabled},
                    reason="导入更新根因条目",
                )
            else:
                row = RootCauseKb(
                    id=uuid4().hex,
                    category_lv1=category_lv1,
                    category_lv2=category_lv2,
                    level=level,
                    content=content_raw,
                    keywords=keywords_raw,
                    is_enabled=is_enabled,
                    created_by=actor.id,
                    updated_by=actor.id,
                )
                db.add(row)
                create_audit_log(
                    db,
                    entity_type="root_cause_kb",
                    entity_id=row.id,
                    action="ROOT_CAUSE_KB_IMPORT_CREATE",
                    actor=actor,
                    after={
                        "categoryLv1": category_lv1,
                        "categoryLv2": category_lv2,
                        "level": level,
                        "content": content_raw,
                        "keywords": keywords_raw,
                        "isEnabled": is_enabled,
                    },
                    reason="导入新增根因条目",
                )

            success += 1
        except Exception as e:
            failed += 1
            msg = str(e)
            if isinstance(e, AppError):
                msg = e.msg
            errors.append({"row": idx, "error": msg})

    return {"total": total, "success": success, "failed": failed, "errors": errors}


def kb_export_xlsx(
    *,
    db: Session,
    keyword: str | None,
    category_lv1: str | None,
    category_lv2: str | None,
    level: str | None,
    is_enabled: bool | None,
) -> str:
    data = kb_page(
        db=db,
        page=1,
        size=5000,
        keyword=keyword,
        category_lv1=category_lv1,
        category_lv2=category_lv2,
        level=level,
        is_enabled=is_enabled,
    )
    records = data.get("records") or []

    wb = Workbook()
    ws = wb.active
    ws.title = "root_cause_kb"
    ws.append(["一级分类", "二级分类", "根因层级", "根因内容", "关键词", "启用"])
    for r in records:
        ws.append(
            [
                r.get("categoryLv1"),
                r.get("categoryLv2"),
                LEVEL_CODE_TO_NAME.get(r.get("level") or "", r.get("level")),
                r.get("content"),
                r.get("keywords") or "",
                "启用" if r.get("isEnabled") else "禁用",
            ]
        )

    out_dir = "/workspace/backend/.data/exports"
    os.makedirs(out_dir, exist_ok=True)
    ts = _now().strftime("%Y%m%d%H%M%S")
    path = os.path.join(out_dir, f"root_cause_kb_{ts}.xlsx")
    wb.save(path)
    return path

